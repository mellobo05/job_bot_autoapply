"""
Excel Application Tracker
Maintains a running Excel spreadsheet of all jobs found / applied to.
"""

import logging
from datetime import datetime
from pathlib import Path
from typing import Dict, Optional

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

logger = logging.getLogger("tracker")

COLUMNS = [
    "Date",
    "Company",
    "Role / Title",
    "Platform",
    "Match %",
    "Status",
    "JD URL",
    "Applied At",
    "HR / Recruiter",
    "LinkedIn Sent",
    "Email Sent",
    "Response Received",
    "Notes",
]

STATUS_COLORS = {
    "Auto-Applied":     "C6EFCE",   # green
    "Review Sent":      "FFEB9C",   # yellow
    "Skipped":          "F2F2F2",   # grey
    "Interviewing":     "BDD7EE",   # blue
    "Offer":            "E2EFDA",   # light green
    "Rejected":         "FFC7CE",   # red
    "Awaiting":         "FCE4D6",   # orange
}


def _get_or_create_wb(path: str) -> openpyxl.Workbook:
    p = Path(path)
    if p.exists():
        return openpyxl.load_workbook(path)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Applications"
    _style_header(ws)
    # Stats sheet
    ws2 = wb.create_sheet("📊 Stats")
    _init_stats_sheet(ws2)
    wb.save(path)
    return wb


def _style_header(ws):
    """Write and style the header row."""
    header_fill  = PatternFill("solid", fgColor="1A1A2E")
    header_font  = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, col_name in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = center_align

    # Column widths
    widths = [14, 22, 34, 16, 10, 16, 42, 18, 22, 14, 12, 18, 28]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"


def _init_stats_sheet(ws):
    ws["A1"] = "📊 Application Statistics"
    ws["A1"].font = Font(bold=True, size=14, color="1A1A2E")
    labels = [
        ("Total Scanned",    "=COUNTA(Applications!A:A)-1"),
        ("Auto-Applied",     "=COUNTIF(Applications!F:F,\"Auto-Applied\")"),
        ("Sent for Review",  "=COUNTIF(Applications!F:F,\"Review Sent\")"),
        ("Interviewing",     "=COUNTIF(Applications!F:F,\"Interviewing\")"),
        ("Offers",           "=COUNTIF(Applications!F:F,\"Offer\")"),
        ("Response Rate",    '=IFERROR(COUNTIF(Applications!L:L,"Yes")/COUNTIF(Applications!F:F,"Auto-Applied"),0)'),
    ]
    for row, (label, formula) in enumerate(labels, 3):
        ws.cell(row=row, column=1, value=label).font = Font(bold=True, name="Calibri")
        ws.cell(row=row, column=2, value=formula)
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 16


def log_job(job: Dict, tracker_config: dict) -> bool:
    """
    Add or update a job entry in the tracker.

    job dict keys:
      title, company, platform, score, status, url,
      applied_at (optional), hr_name (optional),
      linkedin_sent (bool), email_sent (bool),
      response (str), notes (str)
    """
    path = tracker_config["path"]
    Path(path).parent.mkdir(parents=True, exist_ok=True)

    try:
        wb = _get_or_create_wb(path)
        ws = wb["Applications"]

        # Check if job URL already exists (dedup)
        job_url = job.get("url", "")
        for row in ws.iter_rows(min_row=2, values_only=False):
            if row[6].value == job_url:  # col 7 = URL
                _update_row(row, job)
                wb.save(path)
                return True

        # New row
        next_row = ws.max_row + 1
        values = [
            datetime.now().strftime("%Y-%m-%d %H:%M"),
            job.get("company", ""),
            job.get("title", ""),
            job.get("platform", ""),
            job.get("score", ""),
            job.get("status", ""),
            job.get("url", ""),
            job.get("applied_at", ""),
            job.get("hr_name", ""),
            "Yes" if job.get("linkedin_sent") else "No",
            "Yes" if job.get("email_sent") else "No",
            job.get("response", "Awaiting"),
            job.get("notes", ""),
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=next_row, column=col_idx, value=val)
            cell.font      = Font(name="Calibri", size=10)
            cell.alignment = Alignment(vertical="center", wrap_text=(col_idx == 7))
            # Status color coding
            if col_idx == 6:
                color = STATUS_COLORS.get(str(val), "FFFFFF")
                cell.fill = PatternFill("solid", fgColor=color)

        # Alternating row bg
        if next_row % 2 == 0:
            bg = PatternFill("solid", fgColor="F8F8FC")
            for c in range(1, len(COLUMNS) + 1):
                if c != 6:  # don't overwrite status color
                    ws.cell(row=next_row, column=c).fill = bg

        ws.row_dimensions[next_row].height = 22
        wb.save(path)
        logger.info(f"Tracker updated: {job.get('company')} — {job.get('status')}")
        return True

    except Exception as e:
        logger.error(f"Tracker write error: {e}")
        return False


def _update_row(row, job: Dict):
    """Update status/response columns on existing row."""
    if job.get("status"):
        row[5].value = job["status"]
        color = STATUS_COLORS.get(job["status"], "FFFFFF")
        row[5].fill = PatternFill("solid", fgColor=color)
    if job.get("response"):
        row[11].value = job["response"]
    if job.get("hr_name"):
        row[8].value = job["hr_name"]
    if job.get("linkedin_sent"):
        row[9].value = "Yes"
    if job.get("notes"):
        row[12].value = job["notes"]


def update_status(url: str, status: str, tracker_config: dict, notes: str = ""):
    """Update the status of an existing application by URL."""
    log_job({"url": url, "status": status, "notes": notes}, tracker_config)
